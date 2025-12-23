import pandas as pd
import matplotlib.pyplot as plt
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image

class ExcelReportGenerator:

    def __init__(self):
        self.df = None

        # ----------------------------
        # MAIN WINDOW
        # ----------------------------
        self.root = tk.Tk()
        self.root.title("Excel Report Generator")
        self.root.geometry("600x500")    
        self.root.resizable(False, False)

        # ----------------------------
        # STYLING
        # ----------------------------
        style = ttk.Style()
        style.configure("TButton", font=("Segoe UI", 11), padding=6)
        style.configure("TLabel", font=("Segoe UI", 12))
        style.configure("Header.TLabel", font=("Segoe UI", 16, "bold"))

        # ----------------------------
        # MAIN FRAME
        # ----------------------------
        main_frame = ttk.Frame(self.root, padding=25)
        main_frame.pack(expand=True)

        # TITLE
        ttk.Label(
            main_frame,
            text="Excel Report Generator",
            style="Header.TLabel"
        ).pack(pady=(0, 15))

        # DESCRIPTION
        ttk.Label(
            main_frame,
            text="Upload your CSV file and generate a complete\n"
                 "Excel report with summary stats, pivot tables,\n"
                 "and charts — automatically.",
            justify="center"
        ).pack(pady=(0, 25))

        # BUTTONS
        ttk.Button(main_frame, text="📁 Upload CSV File", width=25, 
                   command=self.load_csv).pack(pady=10)

        ttk.Button(main_frame, text="📊 Generate Excel Report", width=25, 
                   command=self.generate_report).pack(pady=10)

        ttk.Button(main_frame, text="❌ Exit", width=25, 
                   command=self.root.destroy).pack(pady=20)

        self.root.mainloop()

    # -------------------------------------------------------------------------
    def load_csv(self):
        file_path = filedialog.askopenfilename(
            title="Select CSV File",
            filetypes=[("CSV Files", "*.csv")]
        )
        if file_path:
            self.df = pd.read_csv(file_path)
            messagebox.showinfo("Success", "CSV Loaded Successfully!")

    # -------------------------------------------------------------------------
    def generate_report(self):
        if self.df is None:
            messagebox.showerror("Error", "Please upload a CSV first!")
            return

        # Summary statistics
        summary = self.df.describe(include="all")

        # Pivot table (Category vs Quantity)
        try:
            pivot = pd.pivot_table(
                self.df,
                index="Category",
                values="Quantity",
                aggfunc="sum"
            )
        except:
            pivot = None
            messagebox.showwarning("Warning", "Pivot could not be generated!")

        # Save as Excel
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            title="Save Excel Report As"
        )
        if not save_path:
            return

        with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
            self.df.to_excel(writer, sheet_name="Raw Data", index=False)
            summary.to_excel(writer, sheet_name="Summary Stats")
            if pivot is not None:
                pivot.to_excel(writer, sheet_name="Pivot Table")

        # Style the Excel (bold headers)
        self.style_excel(save_path)

        # Add chart
        self.generate_chart(save_path)

        messagebox.showinfo("Success", "Report Generated Successfully!")

    # -------------------------------------------------------------------------
    def style_excel(self, file_path):
        wb = load_workbook(file_path)

        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for cell in ws[1]:
                cell.font = Font(bold=True)

        wb.save(file_path)

    # -------------------------------------------------------------------------
    def generate_chart(self, file_path):
        if "Quantity" not in self.df.columns:
            return

        plt.figure(figsize=(6, 4))
        plt.plot(self.df["Quantity"])
        plt.title("Quantity Trend")
        plt.xlabel("Index")
        plt.ylabel("Quantity")

        chart_path = file_path.replace(".xlsx", "_chart.png")
        plt.savefig(chart_path)
        plt.close()

        wb = load_workbook(file_path)
        ws = wb.create_sheet("Chart")
        img = Image(chart_path)
        ws.add_image(img, "A1")

        wb.save(file_path)

# Run App
ExcelReportGenerator()
